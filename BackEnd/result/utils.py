from django.db.models import Sum, Avg,Prefetch
from .models import  ReportSheet, StudentResult ,ResultBatch ,CharacterBatch
from student.models import Student
from school.models import School , Session , Term
from academics.models import ClassRoom
            
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Border, Side
from io import BytesIO 
from typing import Iterable 
from django.http import HttpRequest, HttpResponse 
from django.utils import timezone 
from openpyxl import Workbook , load_workbook
from openpyxl.styles.protection import Protection
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Protection
from openpyxl.formatting.rule import CellIsRule
from openpyxl.formatting.rule import FormulaRule
from rest_framework.exceptions import ValidationError


def build_student_records_workbook(
    request_data,
    students,
    max_marks ,
    *,
    school_name,
    class_name,
    term_name,
    session_name,
    subject_name,
):
    title = f"{class_name}|{subject_name}|{term_name}|{session_name}"
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel max sheet name length

    # ----------------------------
    # HEADER SECRETS
    # ----------------------------
    ws_secret = wb.create_sheet("secret")
    ws_secret.sheet_state = "veryHidden"
    ws_secret["A1"] = request_data.get( "school_id", "")
    ws_secret["A2"] = request_data.get( "session_id", "")
    ws_secret["A3"] = request_data.get( "term_id", "")
    ws_secret["A4"] = request_data.get( "class_id", "")
    ws_secret["A5"] = request_data.get( "subject_id", "")
    
    # ----------------------------
    # HEADER DESIGN
    # ----------------------------
    ws.merge_cells("A1:F1")
    ws["A1"] = school_name
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")

    ws.merge_cells("A2:F2")
    ws["A2"] = f"{class_name} | {subject_name}"
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A2"].font = Font(size=12, bold=True)

    ws.merge_cells("A3:F3")
    ws["A3"] = f"{term_name} | {session_name}"
    ws["A3"].alignment = Alignment(horizontal="center")
    ws["A3"].font = Font(size=12, bold=True)

    ws.merge_cells("A4:F4")
    ws["A4"] = f"Generated: {timezone.now():%Y-%m-%d %H:%M}"
    ws["A4"].alignment = Alignment(horizontal="center")
    ws["A4"].font = Font(size=11, italic=True, color="555555")

    # ----------------------------
    # TABLE HEADER
    # ----------------------------
    headers = ["S/N", "Student Name", "Registration No", f"CA1 ({max_marks['ca1']})", f"CA2 ({max_marks["ca2"]})", f"EXAM ({max_marks['exam']})"]
    ws.append([])
    ws.append(headers)
    header_row = 6

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor="305496")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        cell.protection = Protection(locked=True)

    ws.freeze_panes = "A7"

    # Column widths
    widths = [5, 25, 18, 10, 10, 10]
    for i, width in enumerate(widths):
        ws.column_dimensions[chr(65 + i)].width = width

    # ----------------------------
    # STUDENT ROWS
    # ----------------------------
    start_row = header_row + 1
    for idx, student in enumerate(students, start=1) :
        row = start_row + idx - 1
        ws.cell(row=row, column=1, value=idx).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=student.full_name() )
        ws.cell(row=row, column=3, value=getattr(student, "admission_number", ""))

        # Locked columns
        for col in (1, 2, 3):
            cell = ws.cell(row=row, column=col)
            cell.protection = Protection(locked=True)
            cell.border = thin_border

        # Editable columns
        for col in (4, 5, 6):
            cell = ws.cell(row=row, column=col)
            cell.protection = Protection(locked=False)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    last_row = start_row + len(students) - 1

    # ----------------------------
    # DATA VALIDATION
    # ----------------------------
    if last_row >= start_row:
        # CA1 & CA2 (0-20)
        # dv_ca = DataValidation(type="whole", operator="between", formula1="0", formula2="20", allow_blank=True)
        # dv_ca.add(f"D{start_row}:E{last_row}")
        # ws.add_data_validation(dv_ca)
        
        dv_ca = DataValidation(
            type="custom",
            formula1='=OR(ISNUMBER(D6),UPPER(D6)="ABS",LOWER(D6))',
            allow_blank=True
        )

        dv_ca.error = f"Only numbers (0–{max_marks['ca1']}) or ABS allowed"
        dv_ca.errorTitle = "Invalid Input"
        dv_ca.add(f"D{start_row}:E{last_row}")
        ws.add_data_validation(dv_ca)

        # # EXAM (0-60)
        # dv_exam = DataValidation(type="whole", operator="between", formula1="0", formula2="60", allow_blank=True)
        # dv_exam.add(f"F{start_row}:F{last_row}")
        # ws.add_data_validation(dv_exam)
        dv_exam = DataValidation(
            type="custom",
            formula1='=OR(ISNUMBER(F6),UPPER(F6)="ABS")',
            allow_blank=True
        )

        dv_exam.error = f"Only numbers (0–{max_marks['exam']}) or ABS allowed"
        dv_exam.errorTitle = "Invalid Input"

        dv_exam.add(f"F{start_row}:F{last_row}")
        ws.add_data_validation(dv_exam)

        ws.auto_filter.ref = f"A6:F{last_row}"

    # ----------------------------
    # CONDITIONAL FORMATTING
    # ----------------------------
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # CA1 or CA2 > 20
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    ws.conditional_formatting.add(
        f"D{start_row}:E{last_row}",
        CellIsRule(
            operator="greaterThan",
            formula=[max_marks['ca1']],
            fill=red_fill
        )
    )
    # ws.conditional_formatting.add(f"D{start_row}:E{last_row}", CellIsRule(operator='greaterThan', formula=['20'], stopIfTrue=True, fill=red_fill))
    # EXAM > 60
    ws.conditional_formatting.add(
        f"F{start_row}:F{last_row}",
        CellIsRule(
            operator="greaterThan",
            formula=[max_marks['exam']],
            fill=red_fill
        )
    )
    gray_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    ws.conditional_formatting.add(
        f"D{start_row}:F{last_row}",
        FormulaRule(
            formula=['UPPER(D6)="ABS"'],
            fill=gray_fill
        )
    )
    # ws.conditional_formatting.add(f"F{start_row}:F{last_row}", CellIsRule(operator='greaterThan', formula=['60'], stopIfTrue=True, fill=red_fill))
    # Locked columns altered (1-3)
    
    
    for col_letter in ['A', 'B', 'C']:
        ws.conditional_formatting.add(f"{col_letter}{start_row}:{col_letter}{last_row}",
            CellIsRule(operator='notEqual', formula=[f'{col_letter}{start_row}'], stopIfTrue=True, fill=red_fill)
        )

    # ----------------------------
    # SHEET PROTECTION
    # ----------------------------
    ws.protection.insertRows = False
    ws.protection.deleteRows = False
    ws.protection.insertColumns = False
    ws.protection.deleteColumns = False
    ws.protection.formatCells = False
    ws.protection.sort = False
    ws.protection.autoFilter = True
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = True
    
    # ws.protection.enable()
    # ws.protection.set_password("scores")

    # ----------------------------
    # SAVE TO STREAM
    # ----------------------------
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream

def build_student_character_workbook (
    request_data,
    students,
    *,
    school_name,
    class_name,
    term_name,
    session_name,
    ):
    session = session_name.replace("/","_")
    title = f"{class_name}|{term_name}|{session}"
    print('title: ', title)
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel max sheet name length

    # ----------------------------
    # HEADER SECRETS
    # ----------------------------
    ws_secret = wb.create_sheet("secret")
    ws_secret.sheet_state = "veryHidden"
    ws_secret["A1"] = request_data.get( "school_id", "")
    ws_secret["A2"] = request_data.get( "session_id", "")
    ws_secret["A3"] = request_data.get( "term_id", "")
    ws_secret["A4"] = request_data.get( "class_id", "")
    
    
    # ----------------------------
    # HEADER DESIGN
    # ----------------------------
    ws.merge_cells("A1:K1")
    ws["A1"] = school_name
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    
    ws.merge_cells("A2:K2")
    ws["A2"] = "SKILLS AND PSYCHOMOTOR RECORDS "
    ws["A2"].font = Font(size=12, bold=True, color="FFFFFF")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].fill = PatternFill("solid", fgColor="1F4B78")

    
    ws.merge_cells("A3:K3")
    ws["A3"] = f" {class_name}  {term_name} | {session_name} Students "
    ws["A3"].alignment = Alignment(horizontal="center")
    ws["A3"].font = Font(size=12, bold=True)

    ws.merge_cells("A4:K4")
    ws["A4"] = f"Generated: {timezone.now():%Y-%m-%d %H:%M}"
    ws["A4"].alignment = Alignment(horizontal="center")
    ws["A4"].font = Font(size=11, italic=True, color="555555")

    # ----------------------------  
    # TABLE HEADER
    # ----------------------------
    ws.append([]) # line 4
    ws.merge_cells("A5:K5")
    
    headers = ["S/N", "Student Name", "Registration No","Punctuality", "Honesty","Neatness","Leadership","Handwriting","Verbal Fluency","Creativity"]
    ws.append(headers) # line 5 
    header_row = 6

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor="305496")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        cell.protection = Protection(locked=True)

    ws.freeze_panes = "A7"

    # Column widths
    widths = [8, 25, 18, 12, 12, 12,12,12,12,12,12]
    for i, width in enumerate(widths):
        ws.column_dimensions[chr(65 + i)].width = width

    # ----------------------------
    # STUDENT ROWS
    # ----------------------------
    start_row = header_row + 1
    for idx, student in enumerate(students, start=1) :
        row = start_row + idx - 1
        ws.cell(row=row, column=1, value=idx).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=student.full_name() )
        ws.cell(row=row, column=3, value=getattr(student, "admission_number", ""))

        # Locked columns
        for col in (1, 2, 3):
            cell = ws.cell(row=row, column=col)
            cell.protection = Protection(locked=True)
            cell.border = thin_border

        # Editable columns
        for col in (4, 5, 6,7,8,9,10 ):
            cell = ws.cell(row=row, column=col)
            cell.protection = Protection(locked=False)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    last_row = start_row + len(students) - 1

    # ----------------------------
    # DATA VALIDATION
    # ----------------------------
    if last_row >= start_row:
        # CA1 & CA2 (0-20)
        dv_ca = DataValidation(type="whole", operator="between", formula1="0", formula2="5", allow_blank=True)
        dv_ca.add(f"D{start_row}:K{last_row}")
        ws.add_data_validation(dv_ca)
        ws.auto_filter.ref = f"A6:K{last_row}"

    # ----------------------------
    # CONDITIONAL FORMATTING
    # ----------------------------
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Locked columns altered (1-3)
    for col_letter in ['A', 'B', 'C']:
        ws.conditional_formatting.add(f"{col_letter}{start_row}:{col_letter}{last_row}",
            CellIsRule(operator='notEqual', formula=[f'{col_letter}{start_row}'], stopIfTrue=True, fill=red_fill)
        )

    # ----------------------------
    # SHEET PROTECTION
    # ----------------------------
    ws.protection.insertRows = False
    ws.protection.deleteRows = False
    ws.protection.insertColumns = False
    ws.protection.deleteColumns = False
    ws.protection.formatCells = False
    ws.protection.sort = False
    ws.protection.autoFilter = True
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = True
    
    # ----------------------------
    # SAVE TO STREAM
    # ----------------------------
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream

def decript_scores_from_workbook(workbook_stream,request,students) -> Iterable[dict]:
    wb = load_workbook(workbook_stream,data_only=True)
    ws = wb.active
    try :
        ws_secret = wb["secret"]
    except Exception as e :
        return [], "Invalid sheet"
    school = request.data.get("school")
    session = request.data.get("session")
    term = request.data.get("term")
    classroom = request.data.get("class_room")
    subject = request.data.get("subject")
    
    teacher = request.data.get("teacher")
    
    school_id = ws_secret["A1"].value
    session_id = ws_secret["A2"].value
    term_id = ws_secret["A3"].value
    class_id = ws_secret["A4"].value
    subject_id  = ws_secret["A5"].value
    
    if not all([school_id, session_id, term_id, class_id, subject_id]):
        return [], "missing hidden fields"
    
    def parse_score(value):
        if value is None:
            return 0, False

        value = str(value).strip().upper()

        if value in {"ABS", "A", "AB"}:
            return 0, True

        try:
            return float(value), False
        except (ValueError, TypeError):
            return 0, False
    
    if str(school) != str(school_id):
        return [], "invalid school sheet"
    if str(session) != str(session_id):
        return [], "invalid session sheet"
    if str(term) != str(term_id):
        return [], "invalid term sheet"
    if str(classroom) != str(class_id):
        return [], "invalid class sheet"
    if str(subject) != str(subject_id):
        return [], "invalid subject sheet"

    scores = []
    
    for row in ws.iter_rows(min_row=7, values_only=True):
        if not any(row):
            continue
        student_name = row[1]
        student_id = row[2]  # Assuming admission number is in the 3rd column
        
        ca1, ca1Abs = parse_score(row[3])
        ca2, ca2Abs = parse_score(row[4])
        exam, examAbs = parse_score(row[5])
        
        student_exist = students.filter(admission_number = student_id).first()
        valid_student = student_exist.full_name() == student_name if student_exist else False
        if valid_student :
            scores.append({
                "studentId": student_exist.id ,# we need id not admission number 
                "ca1": ca1,
                "ca2": ca2,
                "exam": exam,
                "ca1Abs": ca1Abs,
                "ca2Abs": ca2Abs,
                "examAbs": examAbs,
            })
    batch_details =  {
        "scores" : scores,
        "teacher": teacher ,
        "class_room" : class_id ,
        "subject" : subject_id ,
        "session" : session_id ,
        "term" : term_id ,
        "school" :school 
    }
    return batch_details , None # no error 

def decript_skills_from_workbook(workbook_stream,request,students) -> Iterable[dict]:
    wb = load_workbook(workbook_stream,data_only=True)
    ws = wb.active
    try :
        ws_secret = wb["secret"]
    except Exception as e :
        return [], "Invalid sheet"
    school = request.data.get("school")
    session = request.data.get("session")
    term = request.data.get("term")
    classroom = request.data.get("class_room")
    
    teacher = request.data.get("teacher")
    
    school_id = ws_secret["A1"].value
    session_id = ws_secret["A2"].value
    term_id = ws_secret["A3"].value
    class_id = ws_secret["A4"].value
    
    if not all([school_id, session_id, term_id, class_id]):
        return [], "missing hidden fields"
    
    if str(school) != str(school_id):
        return [], "school ID mismatch"
    if str(session) != str(session_id):
        return [], "session ID mismatch"
    if str(term) != str(term_id):
        return [], "term ID mismatch"
    if str(classroom) != str(class_id):
        return [], "class ID mismatch"

    charAndScores = []
    for row in ws.iter_rows(min_row=7, values_only=True):
        if not any(row):
            continue
        student_name = row[1]
        student_id = row[2]  # Assuming admission number is in the 3rd column
        punctuality = row[3]
        honesty = row[4]
        netness = row[5]
        leadership = row[6]
        handwriting = row[7]
        verbal_fluency = row[8]
        # verbal_fluency = row[9] 
        creativity = row[9]
        
        student_exist = students.filter(admission_number = student_id).first()
        valid_student = student_exist.full_name() == student_name if student_exist else False
        if valid_student :
            charAndScores.append({
                "studentId": student_exist.id ,# we need id not admission number 
                "punctuality": punctuality,
                "honesty": honesty,
                "neatness": netness,
                "leadership": leadership,
                "handwriting": handwriting,
                "verbal_fluency": verbal_fluency ,
                "creativity": creativity,
            })
            
    details =  {
        "charAndSkills" : charAndScores,
        "teacher": teacher ,
        "class_room" : class_id ,
        "session" : session_id ,
        "term" : term_id ,
        "school" :school 
    }
    return details , None # no error 